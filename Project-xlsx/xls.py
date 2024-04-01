from flask import Flask, render_template
from openpyxl import load_workbook

app = Flask(__name__)

# Route to render HTML template with Excel data
@app.route('/')
def display_excel_data():
    # Read the Excel file
    wb = load_workbook('demo.xlsx')  # Replace 'data.xlsx' with your Excel file
    ws = wb.active
    
    # Extract data from the Excel sheet
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    # Render HTML template with data
    return render_template('excel_data.html', data=data)

@app.route('/hello')
def welcome():
    return 'Hello Anand'

if __name__ == '__main__':
    app.run(debug=True)